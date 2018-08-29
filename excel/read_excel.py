# -*- coding: utf8 -*-
import openpyxl
import re

wb = openpyxl.load_workbook(filename="adc-style-transfer-utters.xlsx")
# sheet_ranges = wb.sheetnames
print(wb.sheetnames)
sheet = wb.worksheets[0]

num_rows = len([1 for row in sheet.rows])
# for row in range(num_rows):
#     print(sheet.cell(column=1, row=row+1).value)

utters = []
for row in sheet.rows:
    # print(row[0].value)
    utters += [row[1].value]

def preprocess(utter):
    ret_utter = utter
    ret_utter = re.sub(r"^/s+", "", ret_utter)
    ret_utter = re.sub(r"[-]+", "", ret_utter)
    ret_utter = re.sub(r"\n+", "", ret_utter)
    return ret_utter

# pre-processing
pp_utters = []
for utter in utters:
    processed = preprocess(utter)
    # print(processed)
    pp_utters += [processed]

# sentence segmentation with simple rule (split by [?!.])
segmented_utters = []
for utter in pp_utters:
    utter = re.sub(r"(:\))+", "", utter)
    utter = re.sub(r"[.]", ".\n", utter)
    utter = re.sub(r"[?]", "?\n", utter)
    utter = re.sub(r"[!]", "!\n", utter)
    splits = re.split(r"\n[\s]?", utter)
    segmented_utters += splits

# remove duplicates and reference sentence (with quotes "")
print("total: %d sentences", len(segmented_utters))
nodup_utters = set(segmented_utters)
print("after removing duplicates: %d sentences", len(nodup_utters))

clean_utters = []
for utter in nodup_utters:
    utter = re.sub(r"^\s+", "", utter)
    utter = re.sub(r"^\)\s+", "", utter)
    if len(utter) < 4:
        continue
    if '"' in utter:
        continue
    clean_utters += [utter]
    print(utter)

print("final cleaned: %d sentences", len(clean_utters))

with open("clean_utters.txt", "w") as wf:
    wf.write('\n'.join(clean_utters).encode("utf8"))
